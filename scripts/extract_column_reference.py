from __future__ import annotations

import json
import os
import re
import sys
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

warnings.filterwarnings(
    'ignore',
    message='Data Validation extension is not supported and will be removed',
    category=UserWarning,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
COLUMN_REFERENCE_PATH = REPO_ROOT / 'src/domain/column/model/column-reference.generated.ts'

WORKBOOK_ENV_VAR = 'COLUMN_REFERENCE_WORKBOOK'
WORKBOOK_CANDIDATES = [
    Path(os.environ[WORKBOOK_ENV_VAR]) if os.environ.get(WORKBOOK_ENV_VAR) else None,
    REPO_ROOT / 'column_calculator_final_release.xlsx',
    REPO_ROOT.parent / 'column_calculator_final_release.xlsx',
    Path('C:/column_calculator_final_release.xlsx'),
]

TS_EXPORT_PATTERN = re.compile(r'export const (\w+) = (.*? as const;)', re.S)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def coerce_float(value: Any) -> float:
    if is_number(value):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if normalized:
            return float(normalized)
    raise ValueError(f'Expected numeric value, got {value!r}')


def load_snapshot_exports(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding='utf-8')
    exports: dict[str, Any] = {}

    for export_name, block in TS_EXPORT_PATTERN.findall(text):
        json_text = block[: -len(' as const;')]
        exports[export_name] = json.loads(json_text)

    if not exports:
        raise ValueError(f'Unable to parse exports from {path}')

    return exports


def find_workbook_path() -> Path | None:
    for candidate in WORKBOOK_CANDIDATES:
        if candidate and candidate.exists():
            return candidate
    return None


def format_numeric_string(value: Any) -> str:
    numeric = coerce_float(value)
    if numeric.is_integer():
        return str(int(numeric))
    return format(numeric, 'g')


def resolve_price_category(profile: str) -> str:
    normalized = profile.strip().lower()
    if normalized.startswith('кв.') or normalized.startswith('пр.'):
        return 'tube'
    return 'i_beam'


def extract_column_brace_unit_mass(ws_calculation) -> float:
    value = ws_calculation['S88'].value
    return coerce_float(value)


def extract_column_candidate_catalog(ws_calculation, expected_count: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []

    for row in range(5, ws_calculation.max_row + 1):
        profile = ws_calculation.cell(row=row, column=12).value
        steel_grade = ws_calculation.cell(row=row, column=10).value
        strength_resistance = ws_calculation.cell(row=row, column=11).value

        if not isinstance(profile, str) or not profile.strip():
            continue
        if not isinstance(steel_grade, str) or not steel_grade.strip():
            continue
        if not is_number(strength_resistance):
            continue

        candidates.append(
            {
                'ordinal': len(candidates) + 1,
                'excluded': ws_calculation.cell(row=row, column=8).value == '-',
                'braceCount': int(coerce_float(ws_calculation.cell(row=row, column=9).value)),
                'priceCategory': resolve_price_category(profile),
                'steelGrade': steel_grade.strip(),
                'strengthResistanceMpa': float(strength_resistance),
                'profile': profile.strip(),
                'areaCm2': coerce_float(ws_calculation.cell(row=row, column=18).value),
                'unitMassKgPerM': coerce_float(ws_calculation.cell(row=row, column=19).value),
                'sectionModulusXCm3': coerce_float(ws_calculation.cell(row=row, column=21).value),
                'radiusXcm': coerce_float(ws_calculation.cell(row=row, column=23).value),
                'radiusYcm': coerce_float(ws_calculation.cell(row=row, column=26).value),
            }
        )

    if len(candidates) != expected_count:
        raise ValueError(
            f'Expected {expected_count} column candidates, extracted {len(candidates)} from {ws_calculation.title!r}'
        )

    return candidates


def extract_column_city_loads(ws_city_loads, expected_count: int) -> list[dict[str, Any]]:
    city_loads: list[dict[str, Any]] = []

    for row in range(3, ws_city_loads.max_row + 1):
        city = ws_city_loads.cell(row=row, column=2).value
        snow_primary = ws_city_loads.cell(row=row, column=3).value
        snow_secondary = ws_city_loads.cell(row=row, column=5).value
        wind_load = ws_city_loads.cell(row=row, column=7).value

        if not isinstance(city, str) or not city.strip():
            continue
        if not is_number(wind_load):
            continue

        snow_value = snow_primary if is_number(snow_primary) else snow_secondary
        if not is_number(snow_value) or not is_number(snow_secondary):
            continue

        city_loads.append(
            {
                'city': city.strip(),
                'snowLoadKpa': float(snow_value),
                'windLoadKpa': float(wind_load),
                'spRegion': format_numeric_string(snow_secondary),
            }
        )

    if len(city_loads) != expected_count:
        raise ValueError(
            f'Expected {expected_count} city loads, extracted {len(city_loads)} from {ws_city_loads.title!r}'
        )

    return city_loads


def extract_column_covering_catalog(ws_summary, expected_count: int) -> list[dict[str, Any]]:
    coverings: list[dict[str, Any]] = []

    for row in range(1, ws_summary.max_row + 1):
        name = ws_summary.cell(row=row, column=12).value
        load = ws_summary.cell(row=row, column=13).value

        if not isinstance(name, str) or not name.strip():
            continue
        if not is_number(load):
            continue

        coverings.append(
            {
                'name': name.strip(),
                'loadKpa': float(load) / 100,
            }
        )

        if len(coverings) == expected_count:
            break

    if len(coverings) != expected_count:
        raise ValueError(
            f'Expected {expected_count} coverings, extracted {len(coverings)} from {ws_summary.title!r}'
        )

    return coverings


def extract_column_support_crane_catalog(ws_cranes, expected_count: int) -> list[dict[str, Any]]:
    cranes: list[dict[str, Any]] = []

    for row in range(2, ws_cranes.max_row + 1):
        capacity_label = ws_cranes.cell(row=row, column=1).value
        span = ws_cranes.cell(row=row, column=2).value
        base = ws_cranes.cell(row=row, column=3).value
        gauge = ws_cranes.cell(row=row, column=4).value
        wheel_load = ws_cranes.cell(row=row, column=5).value
        trolley_mass = ws_cranes.cell(row=row, column=6).value
        crane_mass = ws_cranes.cell(row=row, column=7).value

        if capacity_label is None or not all(
            is_number(value) for value in [span, base, gauge, wheel_load, trolley_mass, crane_mass]
        ):
            continue

        cranes.append(
            {
                'capacityLabel': str(capacity_label).strip(),
                'spanM': int(coerce_float(span)),
                'baseM': coerce_float(base) / 1000,
                'gaugeM': coerce_float(gauge) / 1000,
                'wheelLoadKn': coerce_float(wheel_load),
                'trolleyMassT': coerce_float(trolley_mass),
                'craneMassT': coerce_float(crane_mass),
            }
        )

    if len(cranes) != expected_count:
        raise ValueError(f'Expected {expected_count} cranes, extracted {len(cranes)} from {ws_cranes.title!r}')

    return cranes


def extract_column_exports(workbook_path: Path, snapshot_exports: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    ws_summary = workbook[workbook.sheetnames[4]]
    ws_calculation = workbook[workbook.sheetnames[5]]
    ws_cranes = workbook[workbook.sheetnames[6]]
    ws_city_loads = workbook[workbook.sheetnames[9]]

    return {
        'columnBraceUnitMassKgPerM': extract_column_brace_unit_mass(ws_calculation),
        'columnCandidateCatalog': extract_column_candidate_catalog(
            ws_calculation,
            expected_count=len(snapshot_exports['columnCandidateCatalog']),
        ),
        'columnCityLoads': extract_column_city_loads(
            ws_city_loads,
            expected_count=len(snapshot_exports['columnCityLoads']),
        ),
        'columnCoveringCatalog': extract_column_covering_catalog(
            ws_summary,
            expected_count=len(snapshot_exports['columnCoveringCatalog']),
        ),
        'columnSupportCraneCatalog': extract_column_support_crane_catalog(
            ws_cranes,
            expected_count=len(snapshot_exports['columnSupportCraneCatalog']),
        ),
    }


def render_module(exports_map: dict[str, Any], workbook_path: Path | None, mapped_export_names: list[str]) -> str:
    if workbook_path is None:
        header = [
            '// Rebuilt from the checked-in column reference snapshot.',
            '// Workbook-backed extraction was skipped because no source workbook was found.',
            '',
        ]
    else:
        mapped_list = ', '.join(mapped_export_names)
        header = [
            f'// Rebuilt fully from workbook-backed column references ({workbook_path.name}).',
            f'// Workbook-backed exports: {mapped_list}.',
            '',
        ]

    lines = header
    for export_name, value in exports_map.items():
        serialized = json.dumps(value, ensure_ascii=False, indent=2)
        lines.append(f'export const {export_name} = {serialized} as const;')
        lines.append('')

    return '\n'.join(lines).rstrip() + '\n'


def main() -> int:
    snapshot_exports = load_snapshot_exports(COLUMN_REFERENCE_PATH)
    workbook_path = find_workbook_path()
    mapped_exports: dict[str, Any] = {}

    if workbook_path is not None:
        mapped_exports = extract_column_exports(workbook_path, snapshot_exports)

    final_exports = dict(snapshot_exports)
    final_exports.update(mapped_exports)

    rendered = render_module(final_exports, workbook_path, sorted(mapped_exports.keys()))
    COLUMN_REFERENCE_PATH.write_text(rendered, encoding='utf-8')

    if workbook_path is None:
        print('Rebuilt column references from the checked-in snapshot only (source workbook not found).')
    else:
        print(f'Rebuilt column references from workbook-backed slices ({len(mapped_exports)} exports).')
        print(f'Workbook: {workbook_path}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
