from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
TRUSS_REFERENCE_PATH = REPO_ROOT / 'src/domain/truss/model/truss-reference.generated.ts'

WORKBOOK_ENV_VAR = 'TRUSS_REFERENCE_WORKBOOK'
WORKBOOK_CANDIDATES = [
    Path(os.environ[WORKBOOK_ENV_VAR]) if os.environ.get(WORKBOOK_ENV_VAR) else None,
    REPO_ROOT / 'truss_calculator_molodechno_v1.0.xlsx',
    REPO_ROOT.parent / 'truss_calculator_molodechno_v1.0.xlsx',
]

TS_EXPORT_PATTERN = re.compile(r'export const (\w+) = (.*? as const;)', re.S)

MEMBER_KEYS = [
    'vpN',
    'vpM',
    'vpQ',
    'npNPlus',
    'orbNPlus',
    'orbNMinus',
    'orNPlus',
    'orNMinus',
    'rrNPlus',
    'rrNMinus',
]


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def coerce_float(value: Any) -> float:
    if is_number(value):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if normalized:
            return float(normalized)
    raise ValueError(f'Expected numeric value, got {value!r}')


def load_snapshot_exports(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding='utf-8')
    exports: dict[str, Any] = {}

    for export_name, block in TS_EXPORT_PATTERN.findall(text):
        json_text = block[: -len(' as const;')]
        exports[export_name] = json.loads(json_text)

    if not exports:
        raise ValueError(f'Unable to parse exports from {path}')

    return exports


def find_workbook_path() -> Path | None:
    for candidate in WORKBOOK_CANDIDATES:
        if candidate and candidate.exists():
            return candidate

    downloads = Path('C:/Users/RAN/Downloads')
    if downloads.exists():
        for file in downloads.glob('*.xlsx'):
            if file.name.startswith('~$'):
                continue
            if file.name.endswith('v1.0.xlsx'):
                return file

    return None


def extract_effort_coefficients(ws_epures) -> dict[str, dict[str, dict[str, float]]]:
    component_rows = {
        'snow': [18, 19, 20],
        'covering': [24, 25, 26],
        'windA': [30, 31, 32],
        'windB': [36, 37, 38],
    }

    result: dict[str, dict[str, dict[str, float]]] = {}
    for component, rows in component_rows.items():
        by_span: dict[str, dict[str, float]] = {}
        for row in rows:
            span = int(coerce_float(ws_epures.cell(row=row, column=3).value))
            values: dict[str, float] = {}
            for offset, member_key in enumerate(MEMBER_KEYS):
                values[member_key] = coerce_float(ws_epures.cell(row=row, column=4 + offset).value)
            by_span[str(span)] = values
        result[component] = by_span

    return result


def extract_profile_catalog(ws_vp, ws_np, ws_orb, ws_or, ws_rr) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []

    for row in range(3, 582):
        profile = ws_vp.cell(row=row, column=10).value
        if not isinstance(profile, str) or not profile.strip():
            continue

        catalog.append(
            {
                'ordinal': int(coerce_float(ws_vp.cell(row=row, column=8).value)),
                'profile': profile.strip(),
                'hMm': coerce_float(ws_vp.cell(row=row, column=11).value),
                'bMm': coerce_float(ws_vp.cell(row=row, column=12).value),
                'tMm': coerce_float(ws_vp.cell(row=row, column=13).value),
                'areaCm2': coerce_float(ws_vp.cell(row=row, column=14).value),
                'ixCm4': coerce_float(ws_vp.cell(row=row, column=15).value),
                'wxCm3': coerce_float(ws_vp.cell(row=row, column=16).value),
                'ixRadiusCm': coerce_float(ws_vp.cell(row=row, column=17).value),
                'sxCm3': coerce_float(ws_vp.cell(row=row, column=18).value),
                'iyCm4': coerce_float(ws_vp.cell(row=row, column=19).value),
                'wyCm3': coerce_float(ws_vp.cell(row=row, column=20).value),
                'iyRadiusCm': coerce_float(ws_vp.cell(row=row, column=21).value),
                'unitMassKgPerM': coerce_float(ws_vp.cell(row=row, column=22).value),
                'vpExcluded': ws_vp.cell(row=row, column=7).value == '-',
                'npExcluded': ws_np.cell(row=row, column=6).value == '-',
                'orbExcluded': ws_orb.cell(row=row, column=7).value == '-',
                'orExcluded': ws_or.cell(row=row, column=7).value == '-',
                'rrExcluded': ws_rr.cell(row=row, column=7).value == '-',
            }
        )

    return catalog


def extract_list6_table(ws_list6) -> dict[str, Any]:
    m_axis: list[float] = []
    col = 2
    while True:
        value = ws_list6.cell(row=5, column=col).value
        if value is None:
            break
        m_axis.append(coerce_float(value))
        col += 1

    lambda_axis: list[float] = []
    phi_table: list[list[float]] = []
    row = 6
    while True:
        lambda_value = ws_list6.cell(row=row, column=1).value
        if lambda_value is None:
            break
        lambda_axis.append(coerce_float(lambda_value))

        row_values: list[float] = []
        for offset in range(len(m_axis)):
            row_values.append(coerce_float(ws_list6.cell(row=row, column=2 + offset).value))
        phi_table.append(row_values)
        row += 1

    return {
        'lambdaAxis': lambda_axis,
        'mAxis': m_axis,
        'phiE': phi_table,
    }


def extract_limits_and_constants(ws_main, ws_epures) -> dict[str, Any]:
    return {
        'maxUtilization': {
            'vp': coerce_float(ws_main['D39'].value),
            'np': coerce_float(ws_main['D43'].value),
            'orb': coerce_float(ws_main['D47'].value),
            'or': coerce_float(ws_main['D51'].value),
            'rr': coerce_float(ws_main['D55'].value),
        },
        'minThicknessMm': {
            'vp': coerce_float(ws_main['B22'].value),
            'np': coerce_float(ws_main['B23'].value),
            'orb': coerce_float(ws_main['B24'].value),
            'or': coerce_float(ws_main['B25'].value),
            'rr': coerce_float(ws_main['B26'].value),
        },
        'maxWidthMm': {
            'vp': coerce_float(ws_main['B29'].value),
            'np': coerce_float(ws_main['B30'].value),
        },
        'minWidthMm': {
            'orb': coerce_float(ws_main['B33'].value),
            'or': coerce_float(ws_main['B34'].value),
            'rr': coerce_float(ws_main['B35'].value),
        },
        'extraLoadPercent': coerce_float(ws_epures['B11'].value),
        'massAddConstantKg': 2 * 2 * 4.81,
        'vpBeFactor': 1.12,
        'vpEta': 1.2,
    }


def extract_exports(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True, read_only=False)

    ws_main = wb.worksheets[0]
    ws_epures = wb.worksheets[1]
    ws_vp = wb.worksheets[2]
    ws_list6 = wb.worksheets[3]
    ws_np = wb.worksheets[4]
    ws_orb = wb.worksheets[5]
    ws_or = wb.worksheets[6]
    ws_rr = wb.worksheets[7]

    profile_catalog = extract_profile_catalog(ws_vp, ws_np, ws_orb, ws_or, ws_rr)

    return {
        'trussMemberKeys': MEMBER_KEYS,
        'trussEffortCoefficients': extract_effort_coefficients(ws_epures),
        'trussProfileCatalog': profile_catalog,
        'trussList6Table': extract_list6_table(ws_list6),
        'trussLimitsAndConstants': extract_limits_and_constants(ws_main, ws_epures),
        'trussSnapshotMeta': {
            'sourceWorkbook': workbook_path.name,
            'profileCount': len(profile_catalog),
        },
    }


def render_module(exports_map: dict[str, Any], workbook_path: Path | None) -> str:
    if workbook_path is None:
        header = [
            '// Rebuilt from the checked-in truss reference snapshot.',
            '// Workbook-backed extraction was skipped because no source workbook was found.',
            '',
        ]
    else:
        header = [
            f'// Rebuilt from {workbook_path.name}.',
            '// Source workbook path can be overridden via TRUSS_REFERENCE_WORKBOOK.',
            '',
        ]

    body: list[str] = []
    for export_name, value in exports_map.items():
        body.append(f'export const {export_name} = {json.dumps(value, ensure_ascii=False, indent=2)} as const;')
        body.append('')

    return '\n'.join(header + body).rstrip() + '\n'


def main() -> int:
    workbook_path = find_workbook_path()

    if workbook_path:
        exports_map = extract_exports(workbook_path)
    else:
        if not TRUSS_REFERENCE_PATH.exists():
            raise FileNotFoundError(
                f'Workbook not found and reference file is missing: {TRUSS_REFERENCE_PATH}'
            )
        exports_map = load_snapshot_exports(TRUSS_REFERENCE_PATH)

    TRUSS_REFERENCE_PATH.parent.mkdir(parents=True, exist_ok=True)
    rendered = render_module(exports_map, workbook_path)
    TRUSS_REFERENCE_PATH.write_text(rendered, encoding='utf-8')
    print(f'Wrote {TRUSS_REFERENCE_PATH}')
    return 0


if __name__ == '__main__':
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover - script entrypoint
        print(error, file=sys.stderr)
        raise SystemExit(1)
