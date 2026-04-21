from __future__ import annotations

import json
import math
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

REPO_ROOT = Path(__file__).resolve().parent.parent
PURLIN_REFERENCE_PATH = REPO_ROOT / 'src/domain/purlin/model/purlin-reference.generated.ts'

WORKBOOK_ENV_VAR = 'PURLIN_REFERENCE_WORKBOOK'
WORKBOOK_CANDIDATES = [
    Path(os.environ[WORKBOOK_ENV_VAR]) if os.environ.get(WORKBOOK_ENV_VAR) else None,
    REPO_ROOT / 'calculator_final_release.xlsx',
    REPO_ROOT.parent / 'calculator_final_release.xlsx',
    Path('C:/calculator_final_release.xlsx'),
]

TS_EXPORT_PATTERN = re.compile(r'export const (\w+) = (.*? as const;)', re.S)
ROMAN_REGION_PATTERN = re.compile(r'^[IVX]+[\u0430a]?$')


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def coerce_float(value: Any) -> float:
    if is_number(value):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if normalized:
            return float(normalized)
    raise ValueError(f'Expected numeric value, got {value!r}')



def load_snapshot_exports(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding='utf-8')
    exports: dict[str, Any] = {}

    for export_name, block in TS_EXPORT_PATTERN.findall(text):
        json_text = block[: -len(' as const;')]
        exports[export_name] = json.loads(json_text)

    if not exports:
        raise ValueError(f'Unable to parse exports from {path}')

    return exports



def find_workbook_path() -> Path | None:
    for candidate in WORKBOOK_CANDIDATES:
        if candidate and candidate.exists():
            return candidate
    return None



def extract_wind_region_map(ws) -> dict[str, float]:
    region_map: dict[str, float] = {}

    for column in range(3, 11):
        region = ws.cell(row=1, column=column).value
        load = ws.cell(row=2, column=column).value
        if isinstance(region, str) and is_number(load):
            region_map[region.strip()] = float(load)

    return region_map



def extract_city_loads(ws_city, wind_region_map: dict[str, float], expected_count: int) -> list[dict[str, Any]]:
    city_loads: list[dict[str, Any]] = []

    for row in range(3, 3 + expected_count + 120):
        city = ws_city.cell(row=row, column=2).value
        if not isinstance(city, str) or not city.strip():
            continue

        snow_primary = ws_city.cell(row=row, column=3).value
        snow_secondary = ws_city.cell(row=row, column=5).value
        wind_region = ws_city.cell(row=row, column=6).value

        snow_value = snow_primary if is_number(snow_primary) else snow_secondary
        if not is_number(snow_value) or not isinstance(wind_region, str):
            continue

        wind_load = wind_region_map.get(wind_region.strip())
        if wind_load is None:
            raise ValueError(f'Unsupported wind region {wind_region!r} for city {city!r}')

        city_loads.append(
            {
                'city': city.strip(),
                'snowLoadKpa': float(snow_value),
                'windLoadKpa': wind_load,
            }
        )

        if len(city_loads) == expected_count:
            break

    return city_loads



def extract_profile_sheet_indices(ws_choices, expected_count: int) -> list[dict[str, Any]]:
    profile_sheets: list[dict[str, Any]] = []

    auto_step_index = 1
    for row in range(1, 40):
        profile_sheet = ws_choices.cell(row=row, column=4).value
        if not isinstance(profile_sheet, str):
            continue

        normalized = profile_sheet.strip()
        if not normalized:
            continue
        if not any(character.isdigit() for character in normalized):
            continue
        if ',' not in normalized and '-' not in normalized:
            continue

        profile_sheets.append(
            {
                'profileSheet': normalized,
                'autoStepIndex': auto_step_index,
            }
        )
        auto_step_index += 1

        if len(profile_sheets) == expected_count:
            break

    return profile_sheets



def extract_covering_catalog(ws_summary_legacy, expected_count: int) -> list[dict[str, Any]]:
    coverings: list[dict[str, Any]] = []

    for row in range(1, 140):
        raw_index = ws_summary_legacy[f'K{row}'].value
        raw_name = ws_summary_legacy[f'L{row}'].value
        raw_load = ws_summary_legacy[f'M{row}'].value

        if not is_number(raw_index) or not isinstance(raw_name, str) or not raw_name.strip() or not is_number(raw_load):
            continue

        auto_step_index = int(raw_index)
        coverings.append(
            {
                'name': raw_name.strip(),
                'coveringLoadKpa': float(raw_load) / 100,
                'fixedAutoStepIndex': auto_step_index if auto_step_index >= 5 else None,
            }
        )

        if len(coverings) == expected_count:
            break

    return coverings



def infer_required_panel_thickness(profile_name: str) -> int | None:
    if not profile_name.startswith('2ТПС'):
        return None

    parts = profile_name.split()
    if len(parts) < 2:
        raise ValueError(f'Unable to derive panel thickness from profile {profile_name!r}')

    depth_match = re.match(r'(\d+)', parts[1])
    if not depth_match:
        raise ValueError(f'Unable to derive panel thickness from profile {profile_name!r}')

    depth_mm = int(depth_match.group(1))
    return int(math.ceil(depth_mm / 50) * 50)



def extract_mp_profiles(ws, expected_total_count: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []

    for row in range(5, 5 + expected_total_count + 16):
        profile = ws[f'SK{row}'].value
        moment_resistance = ws[f'SL{row}'].value
        unit_mass = ws[f'SM{row}'].value

        if isinstance(profile, str) and profile.strip() and is_number(moment_resistance) and is_number(unit_mass):
            rows.append(
                {
                    'profile': profile.strip(),
                    'requiredPanelThicknessMm': infer_required_panel_thickness(profile.strip()),
                    'momentResistance': float(moment_resistance),
                    'unitMassKgPerM': float(unit_mass),
                }
            )

            if len(rows) == expected_total_count:
                break

    tps = [item for item in rows if item['profile'].startswith('2ТПС')]
    ps = [item for item in rows if item['profile'].startswith('2ПС')]
    z = [item for item in rows if item['profile'].startswith('Z ')]

    if len(rows) != expected_total_count or not tps or not ps or not z:
        raise ValueError(f'Failed to extract MP profiles on sheet {ws.title!r}')

    return tps, ps, z



def extract_numeric_row(ws, row: int, start_column: str, expected_count: int) -> list[float]:
    values: list[float] = []
    start_index = column_index_from_string(start_column)

    for offset in range(expected_count):
        value = ws.cell(row=row, column=start_index + offset).value
        if not is_number(value):
            raise ValueError(
                f'Expected numeric value at {ws.title}!{start_column}{row}+{offset}, got {value!r}'
            )
        values.append(float(value))

    return values



def extract_mp_step_axis(ws, expected_count: int) -> list[int]:
    return [int(value) for value in extract_numeric_row(ws, row=3, start_column='SO', expected_count=expected_count)]



def extract_auto_step_capacity_table(
    ws_summary_legacy,
    expected_row_count: int,
    expected_capacity_count: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row in range(1, 220):
        step = ws_summary_legacy[f'B{row}'].value
        capacities = [ws_summary_legacy.cell(row=row, column=column).value for column in range(3, 3 + expected_capacity_count)]

        if not is_number(step) or int(step) < 600:
            continue
        if not all(is_number(value) for value in capacities):
            continue

        rows.append(
            {
                'stepMm': int(step),
                'capacityByIndex': [float(value) for value in capacities],
            }
        )

        if len(rows) == expected_row_count:
            break

    return rows



def convert_stability_value(value: Any) -> float:
    if isinstance(value, str) and value.strip().lower() == 'ж':
        return 0.001
    if is_number(value):
        return float(value) / 1000
    raise ValueError(f'Unsupported stability grid value {value!r}')



def extract_sort_stability(ws, expected_lambda_count: int, expected_mef_count: int) -> tuple[list[float], list[float], list[list[float]]]:
    mef_axis = extract_numeric_row(ws, row=8, start_column='B', expected_count=expected_mef_count)
    lambda_axis: list[float] = []
    grid: list[list[float]] = []

    for row in range(9, 9 + expected_lambda_count):
        lambda_value = ws.cell(row=row, column=1).value
        if not is_number(lambda_value):
            raise ValueError(f'Expected lambda value at row {row}, got {lambda_value!r}')

        lambda_axis.append(float(lambda_value))
        grid.append(
            [
                convert_stability_value(ws.cell(row=row, column=column_index_from_string('B') + index).value)
                for index in range(expected_mef_count)
            ]
        )

    return lambda_axis, mef_axis, grid



def extract_wind_height_table(ws, expected_count: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []

    for row in range(52, 52 + expected_count):
        height = ws[f'R{row}'].value
        if not is_number(height):
            raise ValueError(f'Expected wind height at row {row}, got {height!r}')

        result.append(
            {
                'heightM': int(height),
                'kByTerrain': {
                    'А': float(ws[f'K{row}'].value),
                    'В': float(ws[f'L{row}'].value),
                    'С': float(ws[f'M{row}'].value),
                },
                'zetaByTerrain': {
                    'А': float(ws[f'S{row}'].value),
                    'В': float(ws[f'T{row}'].value),
                    'С': float(ws[f'U{row}'].value),
                },
            }
        )

    return result



def extract_nu_table(ws, expected_x_count: int, expected_y_count: int) -> tuple[list[float], list[float], list[list[float]]]:
    y_values = [
        float(ws.cell(row=67, column=column).value)
        for column in range(column_index_from_string('AA'), column_index_from_string('AA') + expected_y_count)
    ]

    x_values: list[float] = []
    grid: list[list[float]] = []
    for row in range(68, 68 + expected_x_count):
        x_value = ws[f'Z{row}'].value
        if not is_number(x_value):
            raise ValueError(f'Expected nu x value at row {row}, got {x_value!r}')
        x_values.append(float(x_value))
        grid.append(
            [
                float(ws.cell(row=row, column=column).value)
                for column in range(column_index_from_string('AA'), column_index_from_string('AA') + expected_y_count)
            ]
        )

    return x_values, y_values, grid



def extract_sort_steel_profiles(ws, snapshot_profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    snapshot_by_ordinal = {int(item['ordinal']): item for item in snapshot_profiles}
    profiles: list[dict[str, Any]] = []

    for row in range(2, 2 + len(snapshot_profiles) + 32):
        ordinal = ws[f'H{row}'].value
        profile_name = ws[f'L{row}'].value

        if not is_number(ordinal) or not isinstance(profile_name, str) or not profile_name.strip():
            if profiles:
                break
            continue

        ordinal_int = int(ordinal)
        snapshot = snapshot_by_ordinal.get(ordinal_int)
        if snapshot is None:
            raise ValueError(f'Missing snapshot metadata for sort steel ordinal {ordinal_int}')

        profiles.append(
            {
                'ordinal': ordinal_int,
                'excluded': bool(snapshot['excluded']),
                'priceCategory': snapshot['priceCategory'],
                'steelGrade': str(ws[f'I{row}'].value).strip(),
                'strengthResistanceMpa': coerce_float(ws[f'J{row}'].value),
                'profile': profile_name.strip(),
                'areaCm2': coerce_float(ws[f'R{row}'].value),
                'unitMassKgPerM': coerce_float(ws[f'S{row}'].value),
                'momentOfInertiaXCm4': coerce_float(ws[f'T{row}'].value),
                'sectionModulusXCm3': coerce_float(ws[f'U{row}'].value),
                'radiusXcm': coerce_float(ws[f'W{row}'].value),
                'momentOfInertiaYCm4': coerce_float(ws[f'X{row}'].value),
                'sectionModulusYCm3': coerce_float(ws[f'Y{row}'].value),
                'radiusYcm': coerce_float(ws[f'Z{row}'].value),
                'eta': coerce_float(ws[f'AA{row}'].value),
                'psi': coerce_float(ws[f'AC{row}'].value),
            }
        )

        if len(profiles) == len(snapshot_profiles):
            break

    return profiles



def extract_sort_steel_step_axis(ws, expected_count: int) -> list[int]:
    return [int(value) for value in extract_numeric_row(ws, row=2, start_column='AG', expected_count=expected_count)]



def extract_sp_rk_en_city_flags(ws, expected_count: int) -> list[dict[str, str]]:
    flags: list[dict[str, str]] = []

    for row in range(2, 2 + expected_count + 80):
        city = ws[f'B{row}'].value
        wind_region = ws[f'C{row}'].value
        if not isinstance(city, str) or not city.strip():
            continue
        if not isinstance(wind_region, str) or not ROMAN_REGION_PATTERN.match(wind_region.strip()):
            continue

        flags.append(
            {
                'city': city.strip(),
                'windRegion': wind_region.strip(),
            }
        )

        if len(flags) == expected_count:
            break

    return flags



def extract_purlin_exports(workbook_path: Path, snapshot_exports: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    ws_summary_legacy = workbook[workbook.sheetnames[7]]
    ws_mp350 = workbook[workbook.sheetnames[8]]
    ws_mp390 = workbook[workbook.sheetnames[9]]
    ws_sort_steel = workbook[workbook.sheetnames[10]]
    ws_stability = workbook[workbook.sheetnames[11]]
    ws_wind_sp = workbook[workbook.sheetnames[12]]
    ws_city_sp20 = workbook[workbook.sheetnames[14]]
    ws_choices = workbook[workbook.sheetnames[16]]
    ws_sp_rk_en = workbook[workbook.sheetnames[18]]

    wind_region_map = extract_wind_region_map(ws_wind_sp)

    mp350_expected_count = (
        len(snapshot_exports['purlinLstkMp3502TpsProfiles'])
        + len(snapshot_exports['purlinLstkMp3502PsProfiles'])
        + len(snapshot_exports['purlinLstkMp350ZProfiles'])
    )
    mp390_expected_count = (
        len(snapshot_exports['purlinLstkMp3902TpsProfiles'])
        + len(snapshot_exports['purlinLstkMp3902PsProfiles'])
        + len(snapshot_exports['purlinLstkMp390ZProfiles'])
    )

    mp350_tps, mp350_ps, mp350_z = extract_mp_profiles(ws_mp350, mp350_expected_count)
    mp390_tps, mp390_ps, mp390_z = extract_mp_profiles(ws_mp390, mp390_expected_count)
    stability_lambda_axis, stability_mef_axis, stability_grid = extract_sort_stability(
        ws_stability,
        expected_lambda_count=len(snapshot_exports['purlinSortSteelStabilityLambdaAxis']),
        expected_mef_count=len(snapshot_exports['purlinSortSteelStabilityMefAxis']),
    )
    nu_x_values, nu_y_values, nu_grid = extract_nu_table(
        ws_wind_sp,
        expected_x_count=len(snapshot_exports['purlinNuXValues']),
        expected_y_count=len(snapshot_exports['purlinNuYValues']),
    )

    return {
        'purlinAutoStepCapacityTable': extract_auto_step_capacity_table(
            ws_summary_legacy,
            expected_row_count=len(snapshot_exports['purlinAutoStepCapacityTable']),
            expected_capacity_count=len(snapshot_exports['purlinAutoStepCapacityTable'][0]['capacityByIndex']),
        ),
        'purlinCityLoads': extract_city_loads(
            ws_city_sp20,
            wind_region_map,
            expected_count=len(snapshot_exports['purlinCityLoads']),
        ),
        'purlinCoveringCatalog': extract_covering_catalog(
            ws_summary_legacy,
            expected_count=len(snapshot_exports['purlinCoveringCatalog']),
        ),
        'purlinLstkMp3502PsProfiles': mp350_ps,
        'purlinLstkMp3502TpsProfiles': mp350_tps,
        'purlinLstkMp350StepAxis': extract_mp_step_axis(
            ws_mp350,
            expected_count=len(snapshot_exports['purlinLstkMp350StepAxis']),
        ),
        'purlinLstkMp350ZProfiles': mp350_z,
        'purlinLstkMp3902PsProfiles': mp390_ps,
        'purlinLstkMp3902TpsProfiles': mp390_tps,
        'purlinLstkMp390ZProfiles': mp390_z,
        'purlinNuGrid': nu_grid,
        'purlinNuXValues': nu_x_values,
        'purlinNuYValues': nu_y_values,
        'purlinProfileSheetIndices': extract_profile_sheet_indices(
            ws_choices,
            expected_count=len(snapshot_exports['purlinProfileSheetIndices']),
        ),
        'purlinSortSteelProfiles': extract_sort_steel_profiles(
            ws_sort_steel,
            snapshot_exports['purlinSortSteelProfiles'],
        ),
        'purlinSortSteelStepAxis': extract_sort_steel_step_axis(
            ws_sort_steel,
            expected_count=len(snapshot_exports['purlinSortSteelStepAxis']),
        ),
        'purlinSortSteelStabilityGrid': stability_grid,
        'purlinSortSteelStabilityLambdaAxis': stability_lambda_axis,
        'purlinSortSteelStabilityMefAxis': stability_mef_axis,
        'purlinSpRkEnCityFlags': extract_sp_rk_en_city_flags(
            ws_sp_rk_en,
            expected_count=len(snapshot_exports['purlinSpRkEnCityFlags']),
        ),
        'purlinWindHeightTable': extract_wind_height_table(
            ws_wind_sp,
            expected_count=len(snapshot_exports['purlinWindHeightTable']),
        ),
    }



def render_module(exports_map: dict[str, Any], workbook_path: Path | None, mapped_export_names: list[str]) -> str:
    if workbook_path is None:
        header = [
            '// Rebuilt from the checked-in purlin reference snapshot.',
            '// Workbook-backed extraction was skipped because no source workbook was found.',
            '',
        ]
    else:
        mapped_list = ', '.join(mapped_export_names)
        if len(mapped_export_names) == len(exports_map):
            header = [
                f'// Rebuilt fully from workbook-backed purlin references ({workbook_path.name}).',
                f'// Workbook-backed exports: {mapped_list}.',
                '',
            ]
        else:
            header = [
                f'// Rebuilt from workbook-backed purlin references ({workbook_path.name}) with snapshot fallback for unmapped slices.',
                f'// Workbook-backed exports: {mapped_list}.',
                '',
            ]

    lines = header
    for export_name, value in exports_map.items():
        serialized = json.dumps(value, ensure_ascii=False, indent=2)
        lines.append(f'export const {export_name} = {serialized} as const;')
        lines.append('')

    return '\n'.join(lines).rstrip() + '\n'



def main() -> int:
    snapshot_exports = load_snapshot_exports(PURLIN_REFERENCE_PATH)
    workbook_path = find_workbook_path()
    mapped_exports: dict[str, Any] = {}

    if workbook_path is not None:
        mapped_exports = extract_purlin_exports(workbook_path, snapshot_exports)

    final_exports = dict(snapshot_exports)
    final_exports.update(mapped_exports)

    rendered = render_module(final_exports, workbook_path, sorted(mapped_exports.keys()))
    PURLIN_REFERENCE_PATH.write_text(rendered, encoding='utf-8')

    if workbook_path is None:
        print('Rebuilt purlin references from the checked-in snapshot only (source workbook not found).')
    else:
        print(
            'Rebuilt purlin references from workbook-backed slices '
            f'({len(mapped_exports)} exports) with snapshot fallback for {len(snapshot_exports) - len(mapped_exports)} exports.'
        )
        print(f'Workbook: {workbook_path}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
